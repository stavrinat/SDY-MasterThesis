#!/usr/bin/env python
# coding: utf-8
#reference: https://realpython.com/openpyxl-excel-spreadsheets-python/


import json
import requests
import spacy
import re
import string
import tiger_tokenizer
import tiger_lemmatizer
import tiger_manage_tokens as tmt
import read_input_data as rid
from collections import Counter
from spacy.tokens import Doc
from openpyxl import Workbook


nlpGR = spacy.load("el")
nlpEN = spacy.load("en")

# function that creates and returns a
# dictionary who contains all tokens that were found in the collection of docs 
# and their frequencies of appearance
def create_dictionary_frequencies(doc):
    dictionary = {}
    for token in doc:
        more_similar = tmt.find_if_similar_key_exists(token,dictionary)
        if more_similar == "none":
            dictionary[token.text] = 1
        else:
            dictionary[more_similar] += 1
            
    return dictionary
        
# which_docs contains all tokens and in which document appears each token
def create_which_document(doc):
    which_docs = {}
    for token in doc:
        more_similar = tmt.find_if_similar_key_exists(token,which_docs)
        if more_similar == "none":
            which_docs[token.text] = []
            
    return which_docs

def create_list_of_terms_in_an_answer(dict_of_all_answers):
    split_answers = {}
    for key in dict_of_all_answers:
        list_of_words = dict_of_all_answers[key].text.split(" ")
        num_of_words = len(list_of_words)
        for i in range(num_of_words-1):
            for j in range(i+1,num_of_words):
                if tmt.similar(list_of_words[i],list_of_words[j]) >= 0.80:
                    list_of_words.remove(list_of_words[j])
                    num_of_words = num_of_words - 1
                    break
 
        split_answers[key] = [word for word in list_of_words if word!='' and word!=" "]
        
    return split_answers

# create a spreadsheet of numeric data corresponding
#to the document collection. Each row is a document, and each column represents a
#feature. Thus, a cell in the spreadsheet is a measurement of a feature (corresponding
#to the column) for a document (corresponding to a row).
def create_spreadsheet(data_file,which_docs):
    workbook = Workbook()
    sheet = workbook.active
    data_to_spreadsheet = rid.dict_of_all_answers_only(data_file)
 
    for key in data_to_spreadsheet:
        data_to_tokenize = data_to_spreadsheet[key]
        doc_tokenization = tiger_tokenizer.tokenization_process(data_to_tokenize)
        doc_lemmatization = tiger_lemmatizer.lemmatization_process(doc_tokenization)
        data_to_spreadsheet[key] = doc_lemmatization

    #split_answers = create_list_of_terms_in_an_answer(data_to_spreadsheet)
    
    # find in which answers exists every token and add answer's 
    # number to the list of corresponding token that was
    # found in answer, in which_docs
    for key in data_to_spreadsheet:
        for word in data_to_spreadsheet[key]:
            if word.text in which_docs.keys() and key not in which_docs[word.text]:
                which_docs[word.text].append(key)
            else:
                more_similar = tmt.find_if_similar_key_exists(word,which_docs)
                if more_similar != "none" and key not in which_docs[more_similar]:
                    which_docs[more_similar].append(key)
    
   # create spreadsheet
    sheet["A1"] = "Answer Number/ID"
    sheet["B1"] = "Text"
    # in first row we will insert the tokens that exist in all answers
    # tokens will be started from column 3 and end up in column = len(which_docs)+3
    col = 3
    for key in which_docs:
        sheet.cell(row=1,column=col).value = key
        col += 1
        
    for r in range(2,len(data_to_spreadsheet)+3):
        sheet.cell(row=r,column=1).value = r-1
        # the below row is in comments because for some reason
        # it doesn't work
        # until further clarification , it shall remain in comments
       # sheet.cell(row=r,column=2).value = tmt.doc_to_string(data_to_spreadsheet[key])
        for c in range(3,len(which_docs)+4):
            sheet.cell(row=r, column=c).value = 0
    
    # i=3 because tokens start from third column
    i =3
    for key in which_docs:
        for j in range(len(which_docs[key])):
            r = which_docs[key][j]+1
            c = i
            sheet.cell(row=r, column=c).value = 1
        i += 1
    workbook.save(filename="presence_of_dictionary_words.xlsx")
    #print(which_docs)
    
    
    